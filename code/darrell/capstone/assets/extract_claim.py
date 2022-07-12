from openpyxl import Workbook, load_workbook

wb = load_workbook('OutlookToExcel.xlsx')
ws = wb.active
ws['B2'].value = "Test"
print(ws['B2'].value)

wb.save('OutlookToExcel.xlsx')
